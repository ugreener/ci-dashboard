"""
Flask web server for dashboard
"""

from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime, timedelta, timezone
from pathlib import Path
import yaml
import threading
import sys
import os
import logging
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from storage.database import DashboardDatabase
from metrics.calculator import MetricsCalculator
from reports.weekly_report import WeeklyReportGenerator

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Global collection status
collection_status = {
    'running': False,
    'progress': '',
    'error': None,
    'completed_at': None,
    'lock': threading.Lock()
}


def _md_cell(val):
    """Escape a value for use in a markdown table cell."""
    return str(val or '').replace('|', '\\|').replace('\n', ' ').replace('\r', ' ').strip()


def _fbc_short(fbc_image):
    """Extract a short display label from an FBC catalog image reference."""
    if not fbc_image:
        return ''
    if '@' in fbc_image:
        return fbc_image.split('@')[-1][:15]
    if ':' in fbc_image:
        return fbc_image.split(':')[-1][:10]
    return fbc_image


def _format_export_row(row, empty_placeholder='-'):
    """Shared row formatting for XLSX/CSV/Markdown exports."""
    job_name = row.get('periodic_job') or ''
    build_id = row.get('build_id') or ''
    step_name = row.get('step_name') or ''
    urls = _build_log_urls(job_name, build_id, step_name)
    short_job = job_name.replace('periodic-ci-medik8s-system-tests-main-', '')
    dur_secs = row.get('job_duration')
    if dur_secs and dur_secs > 0:
        h = int(dur_secs) // 3600
        m = (int(dur_secs) % 3600) // 60
        duration_str = f"{h}h {m}m" if h > 0 else f"{m}m"
    else:
        duration_str = empty_placeholder
    run_date_raw = row.get('run_date') or ''
    run_date = run_date_raw.split('T')[0] if run_date_raw else empty_placeholder
    result = row.get('result') or ''
    result_str = 'PASSED' if result == 'passed' else 'FAILED' if result == 'failed' else (result.upper() or '-')
    return {
        'short_job': short_job, 'duration_str': duration_str,
        'run_date': run_date, 'result_str': result_str, **urls,
    }


GCS_BUCKET = 'test-platform-results'
GCSWEB_HOST = 'gcsweb-ci.apps.ci.l2s4.p1.openshiftapps.com'


def _build_log_urls(job_name, build_id, step_name, gcs_prefix=None):
    """Build GCS and gcsweb log URLs from job metadata."""
    has_job = bool(job_name and build_id)
    if gcs_prefix:
        gcs_base = f"https://storage.googleapis.com/{GCS_BUCKET}/{gcs_prefix}"
        gcsweb_base = f"https://{GCSWEB_HOST}/gcs/{GCS_BUCKET}/{gcs_prefix}"
    else:
        gcs_base = f"https://storage.googleapis.com/{GCS_BUCKET}/logs/{job_name}/{build_id}" if has_job else ''
        gcsweb_base = f"https://{GCSWEB_HOST}/gcs/{GCS_BUCKET}/logs/{job_name}/{build_id}" if has_job else ''
    artifacts_base = f"{gcs_base}/artifacts/{step_name}" if (has_job and step_name) else ''
    return {
        'e2e_log_url': f"{artifacts_base}/e2e-test/build-log.txt" if (has_job and step_name) else '',
        'install_log_url': f"{artifacts_base}/ipi-install-install/build-log.txt" if (has_job and step_name) else '',
        'subscribe_log_url': f"{artifacts_base}/medik8s-operator-subscribe/build-log.txt" if (has_job and step_name) else '',
        'catalog_log_url': f"{artifacts_base}/medik8s-catalogsource/build-log.txt" if (has_job and step_name) else '',
        'artifacts_url': f"{gcsweb_base}/artifacts/{step_name}/gather-must-gather/" if (has_job and step_name) else '',
        'build_log_url': f"{gcs_base}/build-log.txt" if has_job else '',
    }


def run_collection_background(db_path: str, config_file: str = 'config.yaml', days: int = 30):
    """Run data collection in background thread"""
    global collection_status

    try:
        logger.info(f"Starting data collection for {days} days")
        collection_status['progress'] = 'Starting collection...'

        # Load config
        with open(config_file, 'r') as f:
            config = yaml.safe_load(f)

        # Import collector modules
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

        # Initialize collector based on type
        collector_type = config['collector']['type']
        logger.info(f"Using collector type: {collector_type}")

        if collector_type == 'reportportal':
            from collectors.reportportal import ReportPortalCollector
            rp_config = config['collector']['reportportal']
            collector = ReportPortalCollector(rp_config)
        elif collector_type == 'prow_mcp':
            from collectors.prow_mcp import ProwMCPCollector
            mcp_config = config['collector']['prow_mcp']
            collector = ProwMCPCollector(mcp_config)
        elif collector_type == 'prow_gcs':
            from collectors.prow_gcs import ProwGCSCollector
            gcs_config = config['collector']['prow_gcs']
            try:
                collector = ProwGCSCollector(gcs_config)
            except Exception as e:
                error_msg = f'Failed to initialize prow_gcs collector: {e}'
                logger.error(error_msg)
                collection_status['error'] = error_msg
                collection_status['running'] = False
                return
        else:
            error_msg = f'Unsupported collector type: {collector_type}'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Health check
        logger.info("Running health check...")
        collection_status['progress'] = 'Checking data source...'
        if not collector.health_check():
            error_msg = 'Failed to connect to data source'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        # Get job patterns based on collector type
        versions = config['tracking']['versions']
        platforms = config['tracking']['platforms']

        if collector_type == 'reportportal':
            job_patterns = config['collector']['reportportal']['job_patterns']
            # Expand patterns with version placeholders
            expanded_patterns = []
            for pattern in job_patterns:
                for version in versions:
                    expanded_patterns.append(pattern.replace('{version}', version))
        elif collector_type == 'prow_gcs':
            # prow_gcs uses wildcard patterns, no version expansion needed
            # Support both 'job_patterns' (new) and 'job_names' (legacy)
            prow_gcs_config = config['collector']['prow_gcs']
            expanded_patterns = prow_gcs_config.get('job_patterns') or prow_gcs_config.get('job_names', [])
        elif collector_type == 'prow_mcp':
            # prow_mcp uses exact job names from config
            expanded_patterns = None  # Will use job_names from collector config
        else:
            expanded_patterns = []

        # Collect job runs
        logger.info("Collecting job runs...")
        collection_status['progress'] = 'Collecting job runs...'
        job_runs = collector.collect_job_runs(
            start_date=start_date,
            end_date=end_date,
            job_patterns=expanded_patterns,
            versions=versions,
            platforms=platforms
        )
        logger.info(f"Collected {len(job_runs)} job runs")

        # Collect test results
        collection_status['progress'] = f'Collected {len(job_runs)} job runs, collecting test results...'
        logger.info("Collecting test results (fetching logs for failed tests)...")
        test_results = collector.collect_test_results(
            start_date=start_date,
            end_date=end_date,
            job_patterns=expanded_patterns,
            versions=versions,
            platforms=platforms
        )
        logger.info(f"Collected {len(test_results)} test results")

        # Collect presubmit jobs (if configured)
        presubmit_patterns = config.get('collector', {}).get(
            'prow_gcs', {}
        ).get('presubmit_job_patterns', [])
        if presubmit_patterns and collector_type == 'prow_gcs':
            collection_status['progress'] = 'Collecting presubmit job runs...'
            logger.info("Collecting presubmit job runs...")
            presubmit_job_runs = collector.collect_presubmit_job_runs(
                start_date=start_date,
                end_date=end_date,
                job_patterns=presubmit_patterns,
                versions=versions,
                platforms=platforms,
            )
            logger.info(f"Collected {len(presubmit_job_runs)} presubmit job runs")

            collection_status['progress'] = (
                f'Collected {len(presubmit_job_runs)} presubmit job runs, '
                'collecting presubmit test results...'
            )
            logger.info("Collecting presubmit test results...")
            presubmit_test_results = collector.collect_presubmit_test_results(
                start_date=start_date,
                end_date=end_date,
                job_patterns=presubmit_patterns,
                versions=versions,
                platforms=platforms,
                job_runs=presubmit_job_runs,
            )
            logger.info(f"Collected {len(presubmit_test_results)} presubmit test results")

            job_runs.extend(presubmit_job_runs)
            test_results.extend(presubmit_test_results)

        # Save to database
        collection_status['progress'] = (
            f'Collected {len(job_runs)} job runs and {len(test_results)} test results, '
            'saving to database...'
        )
        logger.info("Saving to database...")
        db = DashboardDatabase(db_path)

        inserted_jobs = db.insert_job_runs(job_runs)
        inserted_tests = db.insert_test_results(test_results)

        # Update job_runs with actual test counts from test_results
        logger.info("Updating job runs with test counts...")
        db.conn.execute("""
            UPDATE job_runs
            SET
                total_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status != 'skipped'
                ),
                passed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'passed'
                ),
                failed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'failed'
                ),
                skipped_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'skipped'
                )
            WHERE EXISTS (
                SELECT 1 FROM test_results
                WHERE test_results.job_name = job_runs.job_name
                AND test_results.build_id = job_runs.build_id
            )
        """)
        db.conn.commit()
        logger.info("Job runs updated with test counts")

        # Close connection after write
        db.close()

        logger.info(f"Collection complete! Inserted {inserted_jobs} job runs and {inserted_tests} test results")
        collection_status['progress'] = f'Complete! Saved {inserted_jobs} job runs and {inserted_tests} test results'
        collection_status['error'] = None
        collection_status['completed_at'] = datetime.now(timezone.utc).isoformat()

    except Exception as e:
        logger.error(f"Collection failed: {e}", exc_info=True)
        collection_status['error'] = str(e)
        collection_status['progress'] = 'Failed'
        collection_status['completed_at'] = None
    finally:
        logger.info("Collection thread finished")
        collection_status['running'] = False


def create_app(db_path: str, config: dict = None, config_file: str = 'config.yaml'):
    """
    Create Flask application

    Args:
        db_path: Path to SQLite database
        config: Optional Flask configuration
        config_file: Path to YAML configuration file

    Returns:
        Flask app instance
    """
    app = Flask(__name__,
                template_folder=str(Path(__file__).parent / 'templates'),
                static_folder=str(Path(__file__).parent / 'static'))

    # Disable template caching for development
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.jinja_env.auto_reload = True
    app.jinja_env.cache = {}

    if config:
        app.config.update(config)

    # Load tracking config for blocklist
    blocklist = []
    try:
        with open(config_file, 'r') as f:
            yaml_config = yaml.safe_load(f)
            blocklist = yaml_config.get('tracking', {}).get('blocklist', [])
    except Exception as e:
        print(f"Warning: Could not load blocklist from config: {e}")

    # Initialize database and calculator
    db = DashboardDatabase(db_path)
    calculator = MetricsCalculator(db, blocklist=blocklist)
    report_generator = WeeklyReportGenerator(db, blocklist=blocklist)

    # Check if AI analysis is enabled (default: False for production safety)
    enable_ai = os.environ.get('ENABLE_AI_ANALYSIS', 'false').lower() == 'true'

    def get_latest_version():
        """
        Get the latest version from database.
        Returns the highest version number (e.g., "4.22" if both "4.21" and "4.22" exist)
        """
        query = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC LIMIT 1"
        result = db.execute_query(query)
        return result[0]['version'] if result else None

    def normalize_version(version):
        """
        Normalize version parameter: if empty/None, return latest version.
        This prevents statistically invalid aggregation across different versions.
        """
        if not version or version == '':
            return get_latest_version()
        return version

    @app.route('/')
    def index():
        """Render main dashboard page"""
        # Check if database needs data collection
        global collection_status

        # Check if database is empty or has no recent data
        try:
            # Query for recent data (last 7 days)
            recent_count = db.execute_query(
                "SELECT COUNT(*) as cnt FROM job_runs WHERE timestamp >= datetime('now', '-7 days')"
            )
            needs_collection = recent_count[0]['cnt'] == 0 if recent_count else True

            # Auto-trigger collection if needed and not already running
            if needs_collection and not collection_status['running']:
                with collection_status['lock']:
                    if not collection_status['running']:
                        collection_status['running'] = True
                        collection_status['progress'] = 'Initializing...'
                        collection_status['error'] = None

                        # Start background thread
                        thread = threading.Thread(
                            target=run_collection_background,
                            args=(db_path, config_file, 30),
                            daemon=True
                        )
                        thread.start()

        except Exception as e:
            print(f"Error checking database status: {e}")

        return render_template('dashboard.html', enable_ai=enable_ai)

    @app.route('/logs')
    def view_logs():
        """Display test logs in a new page"""
        log_content = request.args.get('content', '')
        test_name = request.args.get('test', 'Test Log')

        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>{test_name} - Logs</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: #f8fafc;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    background: #1e40af;
                    color: white;
                    padding: 20px;
                    font-size: 18px;
                    font-weight: 600;
                }}
                .content {{
                    padding: 20px;
                }}
                pre {{
                    background: #1e293b;
                    color: #e2e8f0;
                    padding: 20px;
                    border-radius: 6px;
                    overflow-x: auto;
                    font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                    font-size: 13px;
                    line-height: 1.6;
                    white-space: pre-wrap;
                    word-wrap: break-word;
                }}
                .error {{
                    color: #fca5a5;
                }}
                .info {{
                    color: #93c5fd;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">{test_name}</div>
                <div class="content">
                    <pre>{log_content}</pre>
                </div>
            </div>
        </body>
        </html>
        '''
        return html

    @app.route('/api/collection-status')
    def api_collection_status():
        """Get current collection status"""
        global collection_status
        return jsonify({
            'running': collection_status['running'],
            'progress': collection_status['progress'],
            'error': collection_status['error'],
            'completed_at': collection_status['completed_at']
        })

    @app.route('/api/trigger-collection', methods=['POST'])
    def api_trigger_collection():
        """Manually trigger data collection"""
        global collection_status

        days = request.json.get('days', 30) if request.json else 30

        with collection_status['lock']:
            if collection_status['running']:
                return jsonify({'error': 'Collection already running'}), 409

            collection_status['running'] = True
            collection_status['progress'] = 'Initializing...'
            collection_status['error'] = None
            collection_status['completed_at'] = None

            # Start background thread
            thread = threading.Thread(
                target=run_collection_background,
                args=(db_path, config_file, days),
                daemon=True
            )
            thread.start()

        return jsonify({'status': 'started'})

    @app.route('/api/metadata')
    def api_metadata():
        """Get available versions and platforms from database"""
        query_versions = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC"
        query_platforms = "SELECT DISTINCT platform FROM job_runs ORDER BY platform"

        versions = [row['version'] for row in db.execute_query(query_versions)]
        platforms = [row['platform'] for row in db.execute_query(query_platforms)]

        return jsonify({
            'versions': versions,
            'platforms': platforms
        })

    @app.route('/api/summary')
    def api_summary():
        """Get summary statistics"""
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        stats = calculator.get_summary_stats(days=days, version=version, platform=platform)
        return jsonify(stats)

    @app.route('/api/test-results')
    def api_test_results():
        """Get enriched test results with all 17 spreadsheet columns"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))

        rows = db.get_enriched_test_results(days=days, operator=operator, version=version)

        results = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('periodic_job') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name)

            fbc_image = row.get('fbc_image') or ''
            fbc_short_val = _fbc_short(fbc_image)
            fbc_tag_url = ''
            if fbc_image and 'quay.io/' in fbc_image:
                repo_path = fbc_image.split('quay.io/')[-1].split('@')[0].split(':')[0]
                fbc_tag_url = f"https://quay.io/repository/{repo_path}?tab=tags"

            polarion_id = row.get('polarion_id') or ''
            polarion_url = f"https://polarion.engineering.redhat.com/polarion/#/project/OSE/workitem?id={polarion_id}" if polarion_id else ''

            results.append({
                'test_name': row.get('test_name'),
                'test_description': row.get('test_description'),
                'operator': row.get('operator'),
                'result': row.get('result'),
                'periodic_job': job_name,
                'run_date': row.get('run_date'),
                'job_duration': row.get('job_duration'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'fbc_image': fbc_image,
                'fbc_image_short': fbc_short_val,
                'fbc_image_url': fbc_tag_url,
                'prow_url': row.get('job_url') or '',
                **urls,
                'polarion_id': polarion_id,
                'polarion_url': polarion_url,
                'classification': row.get('manual_classification'),
                'jira_key': row.get('jira_issue_key'),
            })

        return jsonify({'results': results, 'total': len(results)})

    @app.route('/api/operator-stats')
    def api_operator_stats():
        """Get per-operator pass/fail counts"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        stats = db.get_operator_stats(days=days, version=version)
        return jsonify({'operators': stats})

    @app.route('/api/job-runs')
    def api_job_runs():
        """Get job run history with enriched metadata"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_job_run_history(days=days, operator=operator, version=version)

        runs = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name)

            runs.append({
                'job_name': job_name,
                'build_id': build_id,
                'status': row.get('status'),
                'run_date': row.get('run_date'),
                'duration': row.get('duration_seconds'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'fbc_image': row.get('fbc_image'),
                'step_name': step_name,
                'total_tests': row.get('total_tests'),
                'passed_tests': row.get('passed_tests'),
                'failed_tests': row.get('failed_tests'),
                'pass_rate': row.get('pass_rate'),
                'prow_url': row.get('job_url') or '',
                **urls,
            })

        return jsonify({'job_runs': runs})

    @app.route('/api/presubmit-results')
    def api_presubmit_results():
        """Get presubmit test results"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_presubmit_test_results(days=days, operator=operator, version=version)

        results = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name, gcs_prefix=row.get('gcs_prefix'))
            pr_number = row.get('pr_number') or row.get('jr_pr_number')

            results.append({
                'test_name': row.get('test_name'),
                'test_description': row.get('test_description'),
                'operator': row.get('operator'),
                'result': row.get('result'),
                'polarion_id': row.get('polarion_id'),
                'pr_number': pr_number,
                'pr_author': row.get('pr_author'),
                'pr_repo': row.get('pr_repo'),
                'job_name': job_name,
                'build_id': build_id,
                'run_date': row.get('run_date'),
                'duration': row.get('job_duration'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'step_name': step_name,
                'prow_url': row.get('job_url') or '',
                **urls,
            })

        return jsonify({'results': results})

    @app.route('/api/presubmit-job-runs')
    def api_presubmit_job_runs():
        """Get presubmit job run history"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_presubmit_job_runs(days=days, operator=operator, version=version)

        runs = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name, gcs_prefix=row.get('gcs_prefix'))

            runs.append({
                'job_name': job_name,
                'build_id': build_id,
                'status': row.get('status'),
                'run_date': row.get('run_date'),
                'duration': row.get('duration_seconds'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'step_name': step_name,
                'total_tests': row.get('total_tests'),
                'passed_tests': row.get('passed_tests'),
                'failed_tests': row.get('failed_tests'),
                'pass_rate': row.get('pass_rate'),
                'pr_number': row.get('pr_number'),
                'pr_author': row.get('pr_author'),
                'pr_repo': row.get('pr_repo'),
                'prow_url': row.get('job_url') or '',
                **urls,
            })

        return jsonify({'job_runs': runs})

    @app.route('/api/trend')
    def api_trend():
        """Get overall pass rate trend"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')

        trend = calculator.get_overall_trend(
            days=days,
            version=version,
            platform=platform
        )
        return jsonify(trend)

    @app.route('/api/test-rankings')
    def api_test_rankings():
        """Get test rankings (worst performers)"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        limit = request.args.get('limit', 20, type=int)

        rankings = calculator.get_test_rankings(
            days=days,
            version=version,
            platform=platform,
            limit=limit
        )
        return jsonify(rankings)

    @app.route('/api/version-comparison')
    def api_version_comparison():
        """Compare pass rates across versions"""
        days = request.args.get('days', 30, type=int)
        comparison = calculator.get_version_comparison(days=days)
        return jsonify(comparison)

    @app.route('/api/platform-comparison')
    def api_platform_comparison():
        """Compare pass rates across platforms"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))

        comparison = calculator.get_platform_comparison(
            days=days,
            version=version
        )
        return jsonify(comparison)

    @app.route('/api/weekly-report')
    def api_weekly_report():
        """Get weekly platform breakdown report"""
        current_days = request.args.get('current_days', 7, type=int)
        previous_days = request.args.get('previous_days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        top = request.args.get('top', 10, type=int)

        # Get platform comparison
        comparison = report_generator.get_platform_week_over_week(
            current_week_days=current_days,
            previous_week_days=previous_days,
            version=version
        )

        # Get top failing tests
        top_tests = calculator.get_test_rankings(days=current_days, version=version, limit=top)

        # Get overall summary
        summary = calculator.get_summary_stats(days=current_days, version=version)

        return jsonify({
            'comparison': comparison,
            'top_tests': top_tests,
            'summary': summary
        })

    @app.route('/api/platform-tests')
    def api_platform_tests():
        """Get test results for a specific platform"""
        platform = request.args.get('platform')
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))

        if not platform:
            return jsonify({'error': 'Platform parameter is required'}), 400

        # Get test rankings for this platform
        tests = calculator.get_test_rankings(days=days, version=version, platform=platform, limit=100)

        # Get platform-specific summary
        summary = calculator.get_summary_stats(days=days, platform=platform, version=version)

        return jsonify({
            'platform': platform,
            'tests': tests,
            'summary': summary,
            'days': days
        })

    @app.route('/api/test-error-by-platform')
    def api_test_error_by_platform():
        """Get latest error for a specific test on a specific platform"""
        test_name = request.args.get('test_name')
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        days = request.args.get('days', 30, type=int)

        if not test_name or not platform:
            return jsonify({'error': 'test_name and platform parameters are required'}), 400

        # Query for most recent failure on this platform
        query = """
            SELECT
                error_message,
                timestamp,
                job_name,
                build_id,
                job_url,
                platform
            FROM test_results
            WHERE test_name = ?
            AND platform = ?
            AND status = 'failed'
            AND error_message IS NOT NULL
            AND timestamp >= datetime('now', ? || ' days')
        """

        params = [test_name, platform, f'-{days}']

        if version:
            query += " AND version = ?"
            params.append(version)

        query += " ORDER BY timestamp DESC LIMIT 1"

        result = db.execute_query(query, params)

        if result:
            return jsonify(result[0])
        else:
            return jsonify({'error': 'No error found for this test/platform combination'}), 404

    @app.route('/api/get-affected-platforms', methods=['POST'])
    def api_get_affected_platforms():
        """Get all platforms affected by a test failure"""
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        days = data.get('days', 7)

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        platforms = db.get_affected_platforms(test_name, version, days)
        return jsonify({'platforms': platforms})

    @app.route('/api/jira/create', methods=['POST'])
    def api_create_jira():
        """Create or find existing Jira issue for a test failure"""
        from integrations import get_jira_integration

        jira = get_jira_integration()
        if not jira:
            return jsonify({
                'status': 'disabled',
                'message': 'Jira integration not configured. Set JIRA_API_TOKEN environment variable.'
            })

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        # Required fields
        test_name = data.get('test_name')
        version = data.get('version')
        platforms = data.get('platforms', [])

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        # If no platforms provided, use single platform from old API
        if not platforms:
            platform = data.get('platform')
            if platform:
                platforms = [platform]

        # Optional fields
        test_description = data.get('test_description', '')
        error_message = data.get('error_message', '')
        job_url = data.get('job_url', '')
        failure_rate = data.get('failure_rate', 0.0)
        runs = data.get('runs', 0)
        failures = data.get('failures', 0)

        # Check for existing issue first (search by test_name + version only)
        existing_issue = jira.search_existing_issue(test_name, version)
        if existing_issue:
            issue_key = existing_issue.get('key')
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'existing',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Found existing issue: {issue_key}'
            })

        # Create new issue
        issue_key = jira.create_issue(
            test_name=test_name,
            test_description=test_description,
            version=version,
            platforms=platforms,
            error_message=error_message,
            job_url=job_url,
            failure_rate=failure_rate,
            runs=runs,
            failures=failures
        )

        if issue_key:
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'created',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Created new issue: {issue_key}'
            })
        else:
            return jsonify({'error': 'Failed to create Jira issue'}), 500

    @app.route('/api/analyze-failure', methods=['POST'])
    def api_analyze_failure():
        """
        Analyze test failure with AI (hybrid: local Claude Code or Anthropic API)
        """
        from ai.analyzer import HybridFailureAnalyzer

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        # Check if we already have a recent analysis
        days = data.get('days', 7)
        existing_analysis = db.get_ai_analysis(test_name, version, platform, days)
        if existing_analysis and data.get('use_cached', True):
            existing_analysis['cached'] = True
            return jsonify(existing_analysis)

        # Use provided error_message or get from database
        error_message = data.get('error_message')
        log_url = data.get('log_url', '')

        if not error_message:
            # Get test error details from database
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)

            query = """
                SELECT error_message, log_url
                FROM test_results
                WHERE test_name = ?
                AND version = ?
                AND platform = ?
                AND status = 'failed'
                AND timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp DESC
                LIMIT 1
            """

            cursor = db.conn.cursor()
            cursor.execute(query, (test_name, version, platform,
                                   start_date.isoformat(), end_date.isoformat()))
            test_data = cursor.fetchone()

            if not test_data:
                return jsonify({'error': 'No recent failure found for this test'}), 404

            error_message = test_data[0] or 'No error message'
            log_url = test_data[1] or ''

        # Analyze with hybrid approach
        try:
            analyzer = HybridFailureAnalyzer()
            analysis = analyzer.analyze_failure(
                test_name=test_name,
                error_message=error_message,
                log_url=log_url,
                platform=platform,
                version=version
            )

            # Save analysis to database
            db.save_ai_analysis(test_name, version, platform, analysis)

            analysis['cached'] = False
            return jsonify(analysis)

        except Exception as e:
            return jsonify({
                'error': f'Analysis failed: {str(e)}',
                'root_cause': 'Analysis service error',
                'confidence': 0
            }), 500

    @app.route('/api/analysis-stats')
    def api_analysis_stats():
        """Get statistics about AI analyses"""
        stats = db.get_analysis_stats()
        return jsonify(stats)

    @app.route('/api/save-classification', methods=['POST'])
    def api_save_classification():
        """
        Save manual classification for a test failure
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')
        classification = data.get('classification')

        if not all([test_name, version, platform, classification]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform, classification'}), 400

        # Validate classification
        valid_classifications = ['product_bug', 'automation_bug', 'system_issue', 'transient', 'to_investigate']
        if classification not in valid_classifications:
            return jsonify({'error': f'Invalid classification. Must be one of: {", ".join(valid_classifications)}'}), 400

        # Save to database
        rows_updated = db.save_manual_classification(
            test_name=test_name,
            version=version,
            platform=platform,
            classification=classification,
            classified_by='user'
        )

        if rows_updated > 0:
            return jsonify({
                'status': 'success',
                'rows_updated': rows_updated,
                'classification': classification
            })
        else:
            return jsonify({'error': 'No matching test result found to update'}), 404

    @app.route('/api/get-test-data', methods=['POST'])
    def api_get_test_data():
        """
        Get existing data for a test (classification, Jira key, AI analysis)
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        result = {
            'manual_classification': None,
            'jira_issue_key': None,
            'ai_analysis': None
        }

        # Get manual classification and Jira issue from test_results
        cursor = db.conn.cursor()

        # Log query parameters for debugging
        logger.info(f"Fetching test data: test_name={test_name}, version={version}, platform={platform}")

        cursor.execute("""
            SELECT manual_classification, jira_issue_key
            FROM test_results
            WHERE test_name = ?
            AND version = ?
            AND UPPER(platform) = UPPER(?)
            AND status = 'failed'
            ORDER BY timestamp DESC
            LIMIT 1
        """, (test_name, version, platform))

        row = cursor.fetchone()
        if row:
            result['manual_classification'] = row[0]
            result['jira_issue_key'] = row[1]
            logger.info(f"Found test data: classification={row[0]}, jira_key={row[1]}")
        else:
            logger.info(f"No test data found for {test_name}/{version}/{platform}")

        # Get AI analysis
        ai_analysis = db.get_ai_analysis(test_name, version, platform, days=90)
        if ai_analysis:
            result['ai_analysis'] = ai_analysis

        return jsonify(result)

    @app.route('/api/export')
    def api_export():
        """Export test results to XLSX, CSV, or MD format matching reference Google Sheet"""
        export_format = request.args.get('format', 'xlsx')
        days = request.args.get('days', 30, type=int)
        version_param = request.args.get('version')
        version = normalize_version(version_param)

        logger.info(f"[EXPORT] Received: format={export_format}, days={days}, version_param={version_param}, normalized_version={version}")

        enriched_rows = db.get_enriched_test_results(days=days, version=version)
        logger.info(f"[EXPORT] Found {len(enriched_rows)} enriched test results")

        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'dashboard-export-{version}-{days}days-{today}'

        if export_format == 'xlsx':
            return export_to_xlsx_enriched(enriched_rows, filename, version, days)
        elif export_format == 'csv':
            return export_to_csv_enriched(enriched_rows, filename, version, days)
        elif export_format == 'md':
            return export_to_markdown_enriched(enriched_rows, filename, version, days)
        else:
            return jsonify({'error': 'Invalid format. Use xlsx, csv, or md'}), 400

    def export_to_xlsx_enriched(enriched_rows, filename, version, days):
        """Export to Excel matching the reference Google Sheet 17-column structure"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Results'

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        link_font = Font(color='0563C1', underline='single', size=10)
        data_font = Font(size=10)
        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pass_font = Font(color='006100', size=10, bold=True)
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        fail_font = Font(color='9C0006', size=10, bold=True)
        alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        headers = [
            'Test Name', 'Operator', 'Result', 'Periodic Job', 'Run Date',
            'Job Duration', 'OCP Version', 'Platform', 'Operator CSV Version',
            'FBC Catalog Image', 'Prow Job', 'E2E Test Log',
            'Operator Install Log', 'CatalogSource Log', 'Artifacts', 'Build Log', 'Polarion ID'
        ]
        col_widths = [48, 11, 10, 30, 14, 14, 44, 11, 37, 42, 13, 16, 24, 24, 15, 13, 15]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='middle', wrap_text=True)

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = 'A2'

        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        for row_idx, row in enumerate(enriched_rows, 2):
            fmt = _format_export_row(row)
            short_job = fmt['short_job']
            duration_str = fmt['duration_str']
            run_date = fmt['run_date']
            result_str = fmt['result_str']

            fbc_image = row.get('fbc_image') or ''
            fbc_short = _fbc_short(fbc_image)

            polarion_id = row.get('polarion_id') or ''

            is_alt = (row_idx % 2 == 0)

            ws.cell(row=row_idx, column=1, value=row.get('test_description') or row.get('test_name')).font = data_font
            ws.cell(row=row_idx, column=2, value=row.get('operator') or '').font = data_font
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

            result_cell = ws.cell(row=row_idx, column=3, value=result_str)
            result_cell.alignment = Alignment(horizontal='center')
            if result_str == 'PASSED':
                result_cell.fill = pass_fill
                result_cell.font = pass_font
            elif result_str == 'FAILED':
                result_cell.fill = fail_fill
                result_cell.font = fail_font

            ws.cell(row=row_idx, column=4, value=short_job).font = data_font
            ws.cell(row=row_idx, column=5, value=run_date).font = data_font
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=6, value=duration_str).font = data_font
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=7, value=row.get('ocp_version') or row.get('version') or '').font = data_font
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=8, value=(row.get('platform') or '').upper()).font = data_font
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=9, value=row.get('csv_version') or '').font = data_font

            fbc_cell = ws.cell(row=row_idx, column=10, value=fbc_short or '-')
            if fbc_image and 'quay.io/' in fbc_image:
                repo_path = fbc_image.split('quay.io/')[-1].split('@')[0].split(':')[0]
                fbc_url = f"https://quay.io/repository/{repo_path}?tab=tags"
                fbc_cell.hyperlink = fbc_url
                fbc_cell.font = link_font

            prow_url = row.get('job_url') or ''
            prow_cell = ws.cell(row=row_idx, column=11, value='View Job' if prow_url else '-')
            prow_cell.alignment = Alignment(horizontal='center')
            if prow_url:
                prow_cell.hyperlink = prow_url
                prow_cell.font = link_font

            e2e_url = fmt['e2e_log_url']
            e2e_cell = ws.cell(row=row_idx, column=12, value='Test Log' if e2e_url else '-')
            e2e_cell.alignment = Alignment(horizontal='center')
            if e2e_url:
                e2e_cell.hyperlink = e2e_url
                e2e_cell.font = link_font

            sub_url = fmt['subscribe_log_url']
            sub_cell = ws.cell(row=row_idx, column=13, value='Install Log' if sub_url else '-')
            sub_cell.alignment = Alignment(horizontal='center')
            if sub_url:
                sub_cell.hyperlink = sub_url
                sub_cell.font = link_font

            cat_url = fmt['catalog_log_url']
            cat_cell = ws.cell(row=row_idx, column=14, value='Catalog Log' if cat_url else '-')
            cat_cell.alignment = Alignment(horizontal='center')
            if cat_url:
                cat_cell.hyperlink = cat_url
                cat_cell.font = link_font

            art_url = fmt['artifacts_url']
            art_cell = ws.cell(row=row_idx, column=15, value='Artifacts' if art_url else '-')
            art_cell.alignment = Alignment(horizontal='center')
            if art_url:
                art_cell.hyperlink = art_url
                art_cell.font = link_font

            bld_url = fmt['build_log_url']
            bld_cell = ws.cell(row=row_idx, column=16, value='Build Log' if bld_url else '-')
            bld_cell.alignment = Alignment(horizontal='center')
            if bld_url:
                bld_cell.hyperlink = bld_url
                bld_cell.font = link_font

            pol_cell = ws.cell(row=row_idx, column=17, value=polarion_id or '-')
            pol_cell.alignment = Alignment(horizontal='center')
            if polarion_id:
                pol_url = f"https://polarion.engineering.redhat.com/polarion/#/project/OSE/workitem?id={polarion_id}"
                pol_cell.hyperlink = pol_url
                pol_cell.font = link_font

            if is_alt:
                for c in range(1, 18):
                    if c == 3:  # skip Result column to preserve pass/fail coloring
                        continue
                    ws.cell(row=row_idx, column=c).fill = alt_fill

        ws.auto_filter.ref = f"A1:Q{len(enriched_rows) + 1}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )

    def export_to_csv_enriched(enriched_rows, filename, version, days):
        """Export enriched test results to CSV with 17-column structure"""
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            'Test Name', 'Operator', 'Result', 'Periodic Job', 'Run Date',
            'Job Duration', 'OCP Version', 'Platform', 'Operator CSV Version',
            'FBC Catalog Image', 'Prow Job URL', 'E2E Test Log URL',
            'Operator Install Log URL', 'CatalogSource Log URL', 'Artifacts URL', 'Build Log URL', 'Polarion ID'
        ]
        writer.writerow(headers)

        for row in enriched_rows:
            fmt = _format_export_row(row, empty_placeholder='')

            writer.writerow([
                row.get('test_description') or row.get('test_name'),
                row.get('operator') or '',
                fmt['result_str'],
                fmt['short_job'],
                fmt['run_date'],
                fmt['duration_str'],
                row.get('ocp_version') or row.get('version') or '',
                (row.get('platform') or '').upper(),
                row.get('csv_version') or '',
                row.get('fbc_image') or '',
                row.get('job_url') or '',
                fmt['e2e_log_url'],
                fmt['subscribe_log_url'],
                fmt['catalog_log_url'],
                fmt['artifacts_url'],
                fmt['build_log_url'],
                row.get('polarion_id') or '',
            ])

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{filename}.csv'
        )

    def export_to_markdown_enriched(enriched_rows, filename, version, days):
        """Export enriched test results to Markdown (8-column format for readability; full 17 columns via XLSX/CSV)"""
        output = io.StringIO()
        output.write(f'# Dashboard Export - {version}\n\n')
        output.write(f'**Time Range:** {days} days | **Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')

        output.write('| Test Name | Operator | Result | Job | Run Date | OCP Version | Platform | Polarion |\n')
        output.write('|-----------|----------|--------|-----|----------|-------------|----------|----------|\n')

        for row in enriched_rows:
            fmt = _format_export_row(row)
            name = _md_cell(row.get('test_description') or row.get('test_name'))
            prow_url = row.get('job_url') or ''
            job_link = f'[{_md_cell(fmt["short_job"])}]({prow_url})' if prow_url else _md_cell(fmt['short_job'])
            ocp_ver = row.get('ocp_version') or row.get('version') or ''

            output.write(f'| {name} | {_md_cell(row.get("operator"))} | {fmt["result_str"]} | {job_link} | {fmt["run_date"]} | {ocp_ver} | {(row.get("platform") or "").upper()} | {_md_cell(row.get("polarion_id"))} |\n')

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/markdown',
            as_attachment=True,
            download_name=f'{filename}.md'
        )

    @app.teardown_appcontext
    def close_db(error):
        """Close database connection on app shutdown"""
        if error:
            print(f"App error: {error}")

    return app
